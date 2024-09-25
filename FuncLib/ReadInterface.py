from tkinter import *
from tkinter import filedialog
import os
import openpyxl
import re
import numpy as np
from scipy.io import savemat
import FuncLib.ReadMapping as Mapping
##接口定义生成相关###############################################################################################################
class SWCSignal:
    SignalName = ""
    SWC_In_Port = []
    SWC_Out_Port = []
    def __init__(self):
        self.SignalName = ""
        self.SWC_In_Port = []
        self.SWC_Out_Port = []
class EnumType:
    EnumName = ""
    Pair ={}
    def __init__(self):
        self.EnumName = ""
        self.Pair = {}
    def GetValue(self,value):
        return self.EnumName + "." + self.Pair[value]
    def display(self):
        print(self.EnumName)
        print(self.Pair)
class Moulde:
    SWC_Name = ""
    InputPort = []
    OutPutPort = []
    def __init__(self):
        self.SWC_Name = ""
        self.OutPutPort = []
        self.InputPort = []
    def display(self):
        print("in")
        for signal in self.InputPort:
            print(signal)
        print("out")
        for signal in self.OutPutPort:
            print(signal)
    def GetModelID(self):
        Tempdata = {}
        Tempdata["InputPort"] = self.InputPort
        Tempdata["OutPutPort"] = self.OutPutPort
        return Tempdata
def isinstr( Str1, SearchStr):
    """
    判断Str1中是否存在SearchStr
    :param Str1:
    :param SearchStr:
    :return:
    """
    try:
        Str1.index(SearchStr)
        return True
    except ValueError:
        return False

def isStrNumber(str1):
    """
    判断字符串是否是一个数字
    :param str1:
    :return:
    """
    try:
        a = int(str1)
        return True
    except ValueError:
        return False

def GetEnumMap(list):
    Pair = {}
    startflag = 0
    for i in range(0,len(list)):
        message = list[i]
        if message == 'Description':
            startflag = 1
            continue
        if startflag:
            if isStrNumber(message):
                Pair[message] = list[i+1]
    return Pair

def GetSysEnumDefine(sheet,num_rows,EnumDefineCol,DescriptionCol):
    DataTypeMap = {}
    for i in range(2,num_rows+1):
        Datatype = str(sheet.cell(i, EnumDefineCol).value)
        if isinstr(Datatype,"Enum:"):
            Description = str(sheet.cell(i, DescriptionCol).value)
            Datatype = Datatype.split(":")[1]
            Datatype = Datatype.replace(" ","")

            DescriptionList = Description.replace("_x000D_",",")
            DescriptionList = DescriptionList.split(",")
            DescriptionList = [item.replace("\n","") for item in DescriptionList]
            if Datatype in DataTypeMap:
                continue
            else:
                TempData = EnumType()
                TempData.EnumName = Datatype
                TempData.Pair = GetEnumMap(DescriptionList)
                DataTypeMap[Datatype] = TempData
    return DataTypeMap
def GetMould_IO(sheet,num_rows):
    MouldeData = {}
    for i in range(2,num_rows+1):
        SWCOutport =str(sheet.cell(i,4).value)
        SWCOutport = SWCOutport.replace(" ", "")
        if SWCOutport == "":
            VCCcomp = str(sheet.cell(i, 1).value)
            if VCCcomp != "":
                SignalName = str(sheet.cell(i, 6).value)
                SignalName = SignalName.replace(" ", "")
                SWCInport = str(sheet.cell(i, 3).value)
                SWCInportList = SWCInport.split(',')
                for sub in SWCInportList:
                    sub = sub.replace(" ", "")
                    if sub in MouldeData:
                        MouldeData[sub].InputPort.append(SignalName)
                    else:
                        MouldeData[sub] = Moulde()
                        MouldeData[sub].InputPort.append(SignalName)

        else:
            SignalName = str(sheet.cell(i,6).value)
            SignalName = SignalName.replace(" ", "")
            if SWCOutport in MouldeData:
                MouldeData[SWCOutport].OutPutPort.append(SignalName)
            else:
                MouldeData[SWCOutport] = Moulde()
                MouldeData[SWCOutport].OutPutPort.append(SignalName)

            SWCInport = str(sheet.cell(i,3).value)
            SWCInportList = SWCInport.split(',')
            for sub in SWCInportList:
                sub = sub.replace(" ", "")
                if sub in MouldeData:
                    MouldeData[sub].InputPort.append(SignalName)
                else:
                    MouldeData[sub] = Moulde()
                    MouldeData[sub].InputPort.append(SignalName)
    return MouldeData

def GetVeh2FuncInterface(MouldeDict,MouldeList):
    InputPort = []
    OutputPort = []
    for Moulde in MouldeList:
        InputPort = InputPort + MouldeDict[Moulde].InputPort
        OutputPort = OutputPort + MouldeDict[Moulde].OutPutPort
    Result = [x for x in InputPort if x not in OutputPort]
    return Result
def GetSWCIOForMat(MouldeDict,MouldeList):
    TempData = {}
    for Moulde in MouldeList:
        TempData[Moulde] = MouldeDict[Moulde].GetModelID()
    return TempData
def SaveWithMat(Dict):
    savemat('SWC_IO.mat',Dict)

def CONCATENATE(SignalName,sub1,sub2,sub3,sub4,sub5):
    Str1 = SignalName
    if sub1 != '':
        Str1 = Str1 + '.' + sub1
    if sub2 != '':
        Str1 = Str1 + '.' + sub2
    if sub3 != '':
        Str1 = Str1 + '.' + sub3
    if sub4 != '':
        Str1 = Str1 + '.' + sub4
    if sub5 != '':
        Str1 = Str1 + '.' + sub5
    return Str1
def GetInterfaceInformation(Sheet,SheetMaxRow,InputsPorts,IgnoreList,SignalNameCol,sub1Col, sub2Col,
                            sub3Col, sub4Col, sub5Col, DimensionCol, DataTypeCol1,DataTypeCol2, initValueCol):
    TotalList = []
    for i in range(2, SheetMaxRow + 1):

        if str(Sheet.cell(i, SignalNameCol).value) in InputsPorts:
            if str(Sheet.cell(i, SignalNameCol).value) in IgnoreList:
                continue
            else:
                tempList = []
                tempList.append(str(Sheet.cell(i, SignalNameCol).value))  # SignalName
                tempList.append(str(Sheet.cell(i, sub1Col).value))  # sub1
                tempList.append(str(Sheet.cell(i, sub2Col).value))  # sub2
                tempList.append(str(Sheet.cell(i, sub3Col).value))  # sub3
                tempList.append(str(Sheet.cell(i, sub4Col).value))  # sub4
                tempList.append(str(Sheet.cell(i, sub5Col).value))  # sub5
                tempList.append(str(Sheet.cell(i, DimensionCol).value))  # Dimension
                if str(Sheet.cell(i, DataTypeCol1).value) != "":
                    tempList.append(str(Sheet.cell(i, DataTypeCol1).value))  # DataType
                else:
                    tempList.append(str(Sheet.cell(i, DataTypeCol2).value))  # DataType
                tempList.append(str(Sheet.cell(i, initValueCol).value))  # InitValue
                TotalList.append(tempList)
    return TotalList


def CreateDefineFile(TotalList,EnumDefineDict):
    CreateNewFileFlag = 1
    CurrentSignal = TotalList[0][0]
    SignalNameList = []
    SignalDefineList = []
    for i in range(0, len(TotalList)):

        SignalName = TotalList[i][0]
        sub1 = TotalList[i][1]
        sub2 = TotalList[i][2]
        sub3 = TotalList[i][3]
        sub4 = TotalList[i][4]
        sub5 = TotalList[i][5]
        Dimension = TotalList[i][6]
        DataType = TotalList[i][7]
        initValue = TotalList[i][8]
        print(SignalName)
        if CurrentSignal != SignalName:  # 信号名称不一致了
            conntent = r"end" + '\n'
            logfile.writelines(conntent)
            logfile.close()  # 吧上一个文件关了
            CreateNewFileFlag = 1
            CurrentSignal = SignalName
        if CreateNewFileFlag:

            # 创建文件并打开
            logName = 'defineReSimInputDataStructure_' + CurrentSignal + '.m'
            logfile = open(logName, 'w')
            conntent = f"function [{CurrentSignal}] = " + 'defineReSimInputDataStructure_' + CurrentSignal + '()' + '\n'
            logfile.writelines(conntent)
            CreateNewFileFlag = 0

            SignalNameList.append(CurrentSignal)
            SignalDefineFuncName = logName.replace('.m','()')
            SignalDefineList.append(SignalDefineFuncName)

        if Mapping.isinstr(DataType, 'Bus:') == False:
            NameStr = CONCATENATE(SignalName, sub1, sub2, sub3, sub4, sub5)
            Dimension = int(Dimension)
            if Dimension == 1:
                if Mapping.isinstr(DataType, 'Enum'):  # 枚举
                    DataType = DataType.split(":")[1]
                    DataType = DataType.replace(" ", "")

                    try:
                        ValueStr = DataType + "." + EnumDefineDict[DataType].Pair["0"]
                    except KeyError:
                        print(DataType)
                        print(EnumDefineDict[DataType])
                        ValueStr = DataType + "." + EnumDefineDict[DataType].Pair["1"]
                else:
                    if initValue != '':
                        ValueStr = DataType + "(" + initValue + ")"
                    else:
                        ValueStr = DataType + "(" + '0' + ")"
            else:
                if Mapping.isinstr(DataType, 'Enum'):  # 枚举
                    DataType = DataType.split(":")[1]
                    DataType = DataType.replace(" ", "")
                    ValueStr = f"zeros({Dimension},1,{DataType})"
                else:
                    ValueStr = f"zeros({Dimension},1,{DataType})"

            conntent = NameStr + '  =  ' + ValueStr + ';' + '\n'
            logfile.writelines(conntent)
    # 创建主函数
    FileName = "InterfaceDefineGenerateFunc.m"
    MainFile = open(FileName, 'w')
    str1 = SignalNameList[0]

    for i  in  range(1,len(SignalNameList)):
        str1 = str1 + ', '+ SignalNameList[i]


    conntent = f"function [{str1}] = " + 'InterfaceDefineGenerateFunc'  + '()' + '\n'
    MainFile.writelines(conntent)
    for i in  range(0,len(SignalNameList)):
        conntent = f" [{SignalNameList[i]}] = {SignalDefineList[i]} ; " + '\n'
        print(conntent)
        MainFile.writelines(conntent)
    conntent = r"end" + '\n'
    MainFile.writelines(conntent)
    MainFile.close()  # 文件关了
    MatFileName = 'Interface.mat'
    A = {}
    A["data"] = SignalNameList
    savemat(MatFileName, A)
