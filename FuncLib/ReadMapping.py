# 此lib对于接口表和mapping
from tkinter import *
from tkinter import filedialog
import os
import openpyxl
import re
from scipy.io import savemat
from FuncLib.ReadDBC import CanMessagedefine

class Veh2Func:
    Interface_Name = ""
    Mapping_Method = ""
    CanDataMessage = CanMessagedefine()
    Sampletime = []
    Value= []
    def __init__(self):
        self.Interface_Name = ""
        self.BaseType = ""
        self.Mapping_Method = ""
        self.CanDataMessage = CanMessagedefine()
        self.Sampletime = []
        self.Value= []
    def isinstr(self,Str1, SearchStr):
        """
        判断Str1中是否存在SearchStr
        :param Str1:
        :param SearchStr:
        :return:
        """
        try:
            Str1.index(SearchStr)
            return True
        except ValueError:
            return False
    def DataConveration(self,inputValue):
        if self.isinstr(self.BaseType,'int'):
            Value = int(inputValue)
        elif self.isinstr(self.BaseType,'single'):
            Value = float(inputValue)
        else:
            Value = int(inputValue)
        return Value
    # def Update(self,CanMessage):
    def display(self):
        print("InterfaceName： ",self.Interface_Name,"BaseType：",self.BaseType,"MappingMethod：",self.Mapping_Method)
        print("CanSignal:")
        self.CanDataMessage.display()
    def GetSimulinkValue(self):
        if len(self.Sampletime) == len(self.Value):
            return [self.Sampletime,self.Value]
        else:
            print("维度不一致")

####公共函数 ############################################################################################################
def SearchFilePath_xlsxV2():
    """
    选择xlsx文件
    :return: 返回选择的xlsx文件地址
    """
    root = Tk()
    root.withdraw()
    FilePath = filedialog.askopenfilename(title='Select The xlsx file', filetypes=[('xlsx files', '*.xlsx')],
    initialdir=os.getcwd())
    file_path, file_name = os.path.split(FilePath)
    os.chdir(file_path)
    return FilePath,file_path
def SearchFilePath_xlsx():
    """
    选择xlsx文件
    :return: 返回选择的xlsx文件地址
    """
    root = Tk()
    root.withdraw()
    FilePath = filedialog.askopenfilename(title='Select The xlsx file', filetypes=[('xlsx files', '*.xlsx')],
    initialdir=os.getcwd())
    file_path, file_name = os.path.split(FilePath)
    os.chdir(file_path)
    return FilePath
def read_Xlsx(ExcelPath):
    """
    读取xlsx文件
    :param ExcelPath: xlsx文件地址
    :return: xlsx对象，sheet 对象
    """
    workbook = openpyxl.load_workbook(ExcelPath)
    # 选择工作表
    sheet = workbook.active
    return workbook, sheet
def ReadFile(FilePath):
    fileHandler = open(FilePath, "r")
    # Get list of all lines ina  file
    listOfLines = fileHandler.readlines()
    fileHandler.close()
    return listOfLines
def isinstr( Str1, SearchStr):
    """
    判断Str1中是否存在SearchStr
    :param Str1:
    :param SearchStr:
    :return:
    """
    try:
        Str1.index(SearchStr)
        return True
    except ValueError:
        return False
def Veh2FuncInterfaceGenerate(InterfaceSheet,InterfaceMaxRow,SignalDatadict):
    InterfaceDict = {}
    for i in range(2, InterfaceMaxRow + 1):
        InterfaceName = str(InterfaceSheet.cell(i, 2).value)
        ByteBase = str(InterfaceSheet.cell(i, 11).value)
        MessageName = str(InterfaceSheet.cell(i, 15).value)
        SignalName = str(InterfaceSheet.cell(i, 16).value)
        MappingMethod = str(InterfaceSheet.cell(i, 22).value)
        MappingValue = str(InterfaceSheet.cell(i, 23).value)

        TempData = Veh2Func()
        TempData.Interface_Name = InterfaceName
        TempData.BaseType = ByteBase
        TempData.Mapping_Method = MappingMethod
        if isinstr(MappingMethod,"Direct"):
            try:
                TempData.CanDataMessage = SignalDatadict[SignalName]
                InterfaceDict[InterfaceName] = TempData
            except KeyError:
                defaultValue = CanMessagedefine()
                TempData.CanDataMessage = defaultValue
                InterfaceDict[InterfaceName] = TempData
                print(InterfaceName,":  Not Find Mapping Signal In DBC")
        elif isinstr(MappingMethod,"Fixed"):
            defaultValue = CanMessagedefine()
            TempData.CanDataMessage = defaultValue
            InterfaceDict[InterfaceName] = TempData
            print(InterfaceName, ":  Not Mapping FixValue",MappingValue)
        else:
            defaultValue = CanMessagedefine()
            TempData.CanDataMessage = defaultValue
            InterfaceDict[InterfaceName] = TempData
            print(InterfaceName, ":  Not Mapping FixValue", MappingValue)
    return InterfaceDict
